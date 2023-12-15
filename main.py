import csv
import openpyxl

class FormulaireSondageExcel:
    def __init__(self):
        self.reponses = {}

    def poser_question(self, question, options=None):
        print(question)
        if options:
            for i, option in enumerate(options, start=1):
                print(f"{i}. {option}")

            choix = input("Choisissez le numéro correspondant à votre réponse : ")
            choix = int(choix)
            reponse = options[choix - 1] if 1 <= choix <= len(options) else None
        else:
            reponse = input("Votre réponse : ")

        return reponse

    def soumettre_formulaire(self):
        self.reponses["Intentions de quitter le pays"] = self.poser_question("Avez-vous l'intention de quitter le pays après avoir terminé vos études universitaires?", ["Oui", "Non", "Incertain"])
        self.reponses["Tranche d'âge"] = self.poser_question("Quel est votre âge actuel?", ["Moins de 20 ans", "20-24 ans", "25-29 ans", "30-34 ans", "35 ans et plus"])
        self.reponses["Niveau d'études"] = self.poser_question("À quel niveau d'études êtes-vous actuellement?", ["Licence 1", "Licence 2", "Licence 3", "Licence 4", "DUT 1", "DUT 2"])
        self.reponses["Pays vise"] = self.poser_question("Vers quel(s) pays envisagez-vous de vous rendre?", ["Etats-Unis", "Canada", "Royaume-Uni", "Australie", "France"])
        if not self.reponses["Pays vise"]:
            self.reponses["Pays vise"] = input("Saisissez le pays désiré : ")
        self.reponses["Raison du départ"] = self.poser_question("Pourquoi envisagez-vous de quitter le pays?" , ["Opportunités professionnelles", "Recherche académique", "Qualité de vie"])
        self.reponses["Objectif du départ"] = self.poser_question("Envisagez-vous de quitter le pays pour des études supplémentaires ou d'autres raisons?", ["Etudes supplementaires","Raisons professionnelles", "Raisons personnelles"])
        self.reponses["Durée prévue à l'étranger"] = self.poser_question("Si vous envisagez un départ temporaire, quelle est la durée prévue de votre séjour?", ["Moins d'un an", "1-2 ans", "3-5 ans", "Plus de 5 ans"])
        self.reponses["Intention de retour dans le pays d'origine"] = self.poser_question("Avez-vous l'intention de retourner dans votre pays d'origine après votre séjour à l'étranger?", ["Oui", "Non", "Incertain"])

    def enregistrer_csv(self, nom_fichier):
        with open(nom_fichier, 'a', newline='', encoding='utf-8') as fichier_csv:
            writer = csv.writer(fichier_csv)

            # Déplace la position d'écriture à la 3e ligne
            if fichier_csv.tell() == 0:
                # Écrire une ligne vide pour occuper les deux premières lignes
                writer.writerow([])
                writer.writerow([])

                # En-têtes à la 3e ligne
                writer.writerow(["Intentions de quitter le pays", "Tranche d'âge", "Niveau d'études", "Pays vise", "Raison du départ", "Objectif du départ", "Durée prévue à l'étranger", "Intention de retour"])

            # Écrire les réponses dans une nouvelle ligne
            writer.writerow([self.reponses[question] for question in self.reponses])

    def enregistrer_excel(self, nom_fichier):
        # Crée un nouveau classeur
        classeur = openpyxl.Workbook()
        feuille = classeur.active

        # Écrire les en-têtes s'il s'agit d'un nouveau fichier
        if feuille.max_row == 0:
            feuille.append(["Intentions de quitter le pays", "Tranche d'âge", "Niveau d'études", "Pays vise", "Raison du départ", "Objectif du départ", "Durée prévue à l'étranger", "Intention de retour dans le pays d'origine"])

        # Écrire les réponses dans une nouvelle ligne
        feuille.append([self.reponses[question] for question in self.reponses])

        # Sauvegarder le classeur
        classeur.save(nom_fichier)
        classeur.close()

if __name__ == "__main__":
    formulaire = FormulaireSondageExcel()
    formulaire.soumettre_formulaire()

    nom_fichier_csv = "reponses_sondage.csv"
    formulaire.enregistrer_csv(nom_fichier_csv)

   #import csv
import openpyxl

class FormulaireSondageExcel:
    def __init__(self):
        self.reponses = {}

    def poser_question(self, question, options=None):
        print(question)
        if options:
            for i, option in enumerate(options, start=1):
                print(f"{i}. {option}")

            choix = input("Choisissez le numéro correspondant à votre réponse : ")
            choix = int(choix)
            reponse = options[choix - 1] if 1 <= choix <= len(options) else None
        else:
            reponse = input("Votre réponse : ")

        return reponse

    def soumettre_formulaire(self):
        self.reponses["Intentions de quitter le pays"] = self.poser_question("Avez-vous l'intention de quitter le pays après avoir terminé vos études universitaires?", ["Oui", "Non", "Incertain"])
        self.reponses["Tranche d'âge"] = self.poser_question("Quel est votre âge actuel?", ["Moins de 20 ans", "20-24 ans", "25-29 ans", "30-34 ans", "35 ans et plus"])
        self.reponses["Niveau d'études"] = self.poser_question("À quel niveau d'études êtes-vous actuellement?", ["Licence 1", "Licence 2", "Licence 3", "Licence 4", "DUT 1", "DUT 2"])
        self.reponses["Pays vise"] = self.poser_question("Vers quel(s) pays envisagez-vous de vous rendre?", ["États-Unis", "Canada", "Royaume-Uni", "Australie", "France"])
        if not self.reponses["Pays vise"]:
            self.reponses["Pays vise"] = input("Saisissez le pays désiré : ")
        self.reponses["Raison du départ"] = self.poser_question("Pourquoi envisagez-vous de quitter le pays?" , ["Opportunités professionnelles", "Recherche académique", "Qualité de vie"])
        self.reponses["Objectif du départ"] = self.poser_question("Envisagez-vous de quitter le pays pour des études supplémentaires ou d'autres raisons?", ["Etudes supplementaires","Raisons professionnelles", "Raisons personnelles"])
        self.reponses["Durée prévue à l'étranger"] = self.poser_question("Si vous envisagez un départ temporaire, quelle est la durée prévue de votre séjour?", ["Moins d'un an", "1-2 ans", "3-5 ans", "Plus de 5 ans"])
        self.reponses["Intention de retour dans le pays d'origine"] = self.poser_question("Avez-vous l'intention de retourner dans votre pays d'origine après votre séjour à l'étranger?", ["Oui", "Non", "Incertain"])

    def enregistrer_csv(self, nom_fichier):
        with open(nom_fichier, 'a', newline='', encoding='utf-8') as fichier_csv:
            writer = csv.writer(fichier_csv)

        # Déplace la position d'écriture à la 3e ligne
        if fichier_csv.tell() == 0:
            # En-têtes à la 3e ligne
            writer.writerow(["Intentions de quitter le pays", "Tranche d'âge", "Niveau d'études", "Pays vise", "Raison du départ", "Objectif du départ", "Durée prévue à l'étranger", "Intention de retour"])

        # Écrire les réponses dans une nouvelle ligne
        writer.writerow([self.reponses[question] for question in ["Intentions de quitter le pays", "Tranche d'âge", "Niveau d'études", "Pays vise", "Raison du départ", "Objectif du départ", "Durée prévue à l'étranger", "Intention de retour"]])


    def enregistrer_excel(self, nom_fichier):
        # Crée un nouveau classeur
        classeur = openpyxl.Workbook()
        feuille = classeur.active

        # Écrire les en-têtes s'il s'agit d'un nouveau fichier
        if feuille.max_row == 0:
            feuille.append(["Intentions de quitter le pays", "Tranche d'âge", "Niveau d'études", "Pays vise", "Raison du départ", "Objectif du départ", "Durée prévue à l'étranger", "Intention de retour dans le pays d'origine"])

        # Écrire les réponses dans une nouvelle ligne
        feuille.append([self.reponses[question] for question in self.reponses])

        # Sauvegarder le classeur
        classeur.save(nom_fichier)
        classeur.close()

if __name__ == "__main__":
    formulaire = FormulaireSondageExcel()
    formulaire.soumettre_formulaire()

    nom_fichier_csv = "reponses_sondage.csv"
    formulaire.enregistrer_csv(nom_fichier_csv)

    #nom_fichier_excel = "reponses_sondage.xlsx"
    #formulaire.enregistrer_excel(nom_fichier_excel)

    print(f"Les réponses ont été enregistrées dans les fichiers : {nom_fichier_csv}")

